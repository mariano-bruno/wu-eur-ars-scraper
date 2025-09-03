#!/usr/bin/env python3
"""
Scrape Western Union EUR->ARS rate y lo guarda en un Excel (.xlsx)
con columnas: Fecha, Hora, Cotización
"""
import re
import os
from datetime import datetime
try:
    from zoneinfo import ZoneInfo
except Exception:
    ZoneInfo = None

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry
from openpyxl import Workbook, load_workbook

URL = "https://www.westernunion.com/es/es/currency-converter/eur-to-ars-rate.html"
XLSX_PATH = os.environ.get("WU_XLSX_PATH", "wu_eur_ars.xlsx")
TZ = os.environ.get("LOCAL_TZ", "America/Argentina/Cordoba")

def get_session():
    s = requests.Session()
    retries = Retry(
        total=4,
        backoff_factor=0.8,
        status_forcelist=(429, 500, 502, 503, 504),
        allowed_methods=frozenset(["GET"]),
    )
    s.mount("https://", HTTPAdapter(max_retries=retries))
    s.headers.update({
        "User-Agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) "
                      "AppleWebKit/537.36 (KHTML, like Gecko) "
                      "Chrome/124.0.0.0 Safari/537.36",
        "Accept-Language": "es-ES,es;q=0.9,en;q=0.8",
    })
    return s

PATTERNS = [
    r"FX:\s*1(?:[.,]00)?\s*EUR\s*[–\-]\s*([0-9]+(?:[.,][0-9]+)*)\s*ARS",
    r"FX:\s*1(?:[.,]00)?\s*EUR[^0-9]{0,20}([0-9]+(?:[.,][0-9]+)*)\s*ARS",
]

def parse_rate(html: str) -> float:
    for pat in PATTERNS:
        m = re.search(pat, html, re.IGNORECASE)
        if m:
            raw = m.group(1).replace(",", ".")
            return float(raw)
    raise RuntimeError("No pude encontrar la cotización en el HTML.")

def now_local(tzname: str):
    try:
        tz = ZoneInfo(tzname)
        dt = datetime.now(tz)
    except Exception:
        dt = datetime.now()
    return dt.strftime("%Y-%m-%d"), dt.strftime("%H:%M:%S")

def append_xlsx(fecha: str, hora: str, rate: float, path: str):
    if os.path.exists(path):
        wb = load_workbook(path)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "EUR_ARS"
        ws.append(["Fecha", "Hora", "Cotización"])

    ws.append([fecha, hora, round(rate, 6)])
    wb.save(path)

def main():
    s = get_session()
    resp = s.get(URL, timeout=30)
    resp.raise_for_status()
    rate = parse_rate(resp.text)
    fecha, hora = now_local(TZ)
    append_xlsx(fecha, hora, rate, XLSX_PATH)
    print(f"OK {fecha} {hora} rate={rate:.6f} guardado en {XLSX_PATH}")

if __name__ == "__main__":
    main()
